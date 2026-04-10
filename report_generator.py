# -*- coding: utf-8 -*-
"""Generate XLSX and PDF reports from tender search results."""

import os
from typing import List, Dict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter


# ─────────────────────── XLSX ───────────────────────

def generate_xlsx(items: List[Dict], results_map: Dict[int, List[Dict]],
                  tender_name: str, filepath: str):
    """
    items: list of tender_items dicts
    results_map: {item_id: [search_result dicts]}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты поиска"

    # ── Styles ──
    header_fill   = PatternFill("solid", fgColor="1A3A5C")
    header_font   = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    subhdr_fill   = PatternFill("solid", fgColor="2D6DA8")
    subhdr_font   = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
    found_fill    = PatternFill("solid", fgColor="EBF5E8")
    noprice_fill  = PatternFill("solid", fgColor="FFF8E1")
    notfound_fill = PatternFill("solid", fgColor="FDECEA")
    best_fill     = PatternFill("solid", fgColor="D4EDDA")
    link_font     = Font(color="1565C0", underline="single", size=9, name="Calibri")
    wrap_align    = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border        = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ── Title row ──
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"Тендерный поиск: {tender_name}   |   Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    title_cell.font = Font(bold=True, size=12, color="1A3A5C", name="Calibri")
    title_cell.alignment = center_align
    title_cell.fill = PatternFill("solid", fgColor="E8F0F7")
    ws.row_dimensions[1].height = 28

    # ── Header row ──
    HEADERS = [
        ("№",          4),
        ("Наименование позиции", 35),
        ("Характеристики\nиз заявки", 40),
        ("Кол-во",    12),
        ("Макс. цена\n(тендер), р.", 14),
        ("Поставщик",  24),
        ("Найденная\nцена, р.",  14),
        ("Ссылка",    50),
    ]
    for col, (name, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = "A3"

    desc_font  = Font(italic=True, size=8, name="Calibri", color="2D5A8E")
    desc_fill  = PatternFill("solid", fgColor="EEF4FB")

    row = 3
    for item in items:
        item_id = item["id"]
        offers = results_map.get(item_id, [])
        description = item.get("description", "") or ""

        # Columns: A=№, B=Наименование, C=Характеристики, D=Кол-во, E=Макс.цена, F=Поставщик, G=Цена, H=Ссылка

        if not offers:
            for c in range(1, 9):
                cell = ws.cell(row=row, column=c)
                cell.fill = notfound_fill
                cell.alignment = wrap_align
                cell.border = border
            ws.cell(row=row, column=1, value=item["num"])
            ws.cell(row=row, column=2, value=item["name"]).font = Font(bold=True, size=9, name="Calibri")
            ws.cell(row=row, column=3, value=description).font = Font(italic=True, size=8, name="Calibri", color="555555")
            qty_cell_nf = ws.cell(row=row, column=4, value=_fmt_qty_num(item["qty"]))
            qty_cell_nf.alignment = center_align
            mp_cell_nf = ws.cell(row=row, column=5, value=_fmt_price(item["max_price"]) if item.get("max_price") else "—")
            mp_cell_nf.alignment = center_align
            ws.cell(row=row, column=6, value="— не найдено —")
            ws.row_dimensions[row].height = 36
            row += 1
            continue

        n = len(offers)
        start_row = row

        # Merge B, C, D, E across all offer rows for this item
        for col in [2, 3, 4, 5]:
            if n > 1:
                ws.merge_cells(f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{start_row + n - 1}")

        # Name cell
        name_cell = ws.cell(row=start_row, column=2, value=item["name"])
        name_cell.font = Font(bold=True, size=9, name="Calibri")
        name_cell.fill = subhdr_fill
        name_cell.border = border
        name_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Description cell
        desc_cell = ws.cell(row=start_row, column=3, value=description)
        desc_cell.font = desc_font
        desc_cell.fill = desc_fill
        desc_cell.border = border
        desc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Qty cell
        qty_cell = ws.cell(row=start_row, column=4, value=_fmt_qty_num(item["qty"]))
        qty_cell.fill = subhdr_fill
        qty_cell.border = border
        qty_cell.alignment = Alignment(horizontal="center", vertical="top")
        qty_cell.font = Font(size=9, name="Calibri")

        # Max price cell
        mp_cell = ws.cell(row=start_row, column=5, value=_fmt_price(item["max_price"]) if item.get("max_price") else "—")
        mp_cell.fill = subhdr_fill
        mp_cell.border = border
        mp_cell.alignment = Alignment(horizontal="center", vertical="top")
        mp_cell.font = Font(size=9, name="Calibri")

        # Style merged cells in extra rows
        for extra_row in range(start_row + 1, start_row + n):
            for col in [2, 3, 4, 5]:
                cell = ws.cell(row=extra_row, column=col)
                cell.fill = desc_fill if col == 3 else subhdr_fill
                cell.border = border

        # Offer rows
        for i, offer in enumerate(offers):
            r = start_row + i
            price = float(offer.get("price", 0) or 0)
            is_best = (i == 0 and price > 0)
            fill = best_fill if is_best else (found_fill if price > 0 else noprice_fill)

            def _cell(c, v, _r=r, _fill=fill):
                cell = ws.cell(row=_r, column=c, value=v)
                cell.fill = _fill
                cell.alignment = wrap_align
                cell.border = border
                return cell

            num_cell = ws.cell(row=r, column=1, value="★" if is_best else f"{i+1}.")
            num_cell.fill = fill
            num_cell.alignment = center_align
            num_cell.border = border
            if is_best:
                num_cell.font = Font(bold=True, color="1B5E20", size=10, name="Calibri")

            _cell(6, offer.get("supplier", ""))
            pr_cell = _cell(7, price if price else "н/д")
            pr_cell.alignment = center_align
            if is_best:
                pr_cell.font = Font(bold=True, color="1B5E20", size=10, name="Calibri")

            url = offer.get("url", "")
            url_cell = _cell(8, offer.get("title", url)[:70] or url[:70])
            if url:
                url_cell.hyperlink = url
                url_cell.font = link_font

            ws.row_dimensions[r].height = 36

        row = start_row + n
        ws.row_dimensions[row].height = 5
        row += 1

    # Summary
    ws.cell(row=row + 1, column=1, value="ИТОГО позиций:").font = Font(bold=True, name="Calibri")
    ws.cell(row=row + 1, column=2, value=len(items))
    ws.cell(row=row + 2, column=1, value="Найдено:").font = Font(bold=True, name="Calibri")
    ws.cell(row=row + 2, column=2, value=sum(1 for it in items if results_map.get(it["id"])))

    wb.save(filepath)


def _fmt_qty(qty, unit):
    if qty == int(qty):
        return f"{int(qty)} {unit}"
    return f"{qty} {unit}"


def _fmt_qty_num(qty):
    """Return quantity as number only, no unit."""
    if qty == int(qty):
        return int(qty)
    return qty


def _fmt_price(price):
    if not price:
        return "—"
    return f"{price:,.0f}".replace(",", " ")


# ─────────────────────── PDF ───────────────────────

def generate_pdf(items: List[Dict], results_map: Dict[int, List[Dict]],
                 tender_name: str, filepath: str):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    # Register fonts with Cyrillic support
    _register_fonts()

    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(A4),
        rightMargin=10*mm, leftMargin=10*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    styles = getSampleStyleSheet()
    body_font = "DejaVu" if _font_available("DejaVu") else "Helvetica"

    title_style = ParagraphStyle("title", fontName=body_font + "-Bold" if body_font != "Helvetica" else "Helvetica-Bold",
                                  fontSize=13, spaceAfter=6, alignment=TA_CENTER, textColor=colors.HexColor("#1A3A5C"))
    normal_style = ParagraphStyle("normal", fontName=body_font, fontSize=8, leading=10)
    small_style  = ParagraphStyle("small",  fontName=body_font, fontSize=7, leading=9,
                                   textColor=colors.HexColor("#444444"))
    bold_style   = ParagraphStyle("bold",   fontName=body_font + "-Bold" if body_font != "Helvetica" else "Helvetica-Bold",
                                   fontSize=8, leading=10)
    link_style   = ParagraphStyle("link",   fontName=body_font, fontSize=7, leading=9,
                                   textColor=colors.HexColor("#1565C0"))

    COLOR_HEADER  = colors.HexColor("#1A3A5C")
    COLOR_SUBHDR  = colors.HexColor("#2D6DA8")
    COLOR_BEST    = colors.HexColor("#D4EDDA")
    COLOR_FOUND   = colors.HexColor("#EBF5E8")
    COLOR_NOPRICE = colors.HexColor("#FFF8E1")
    COLOR_NOTFOUND = colors.HexColor("#FDECEA")
    WHITE         = colors.white

    story = []

    story.append(Paragraph(f"Тендерный поиск: {tender_name}", title_style))
    story.append(Paragraph(
        f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
        f"Позиций: {len(items)}  |  "
        f"Найдено: {sum(1 for it in items if results_map.get(it['id']))}",
        small_style
    ))
    story.append(Spacer(1, 4*mm))

    # Columns: №, Наименование, Кол-во, Макс.цена, Поставщик, Цена, Ссылка
    col_widths = [8*mm, 70*mm, 16*mm, 20*mm, 38*mm, 18*mm, 80*mm]

    def _hdr(text, align=TA_CENTER):
        fn = body_font + "-Bold" if body_font != "Helvetica" else "Helvetica-Bold"
        return Paragraph(text, ParagraphStyle("h", fontName=fn, fontSize=8, textColor=WHITE, alignment=align))

    header_row = [
        _hdr("№"),
        _hdr("Наименование", TA_LEFT),
        _hdr("Кол-во"),
        _hdr("Макс. цена, р."),
        _hdr("Поставщик", TA_LEFT),
        _hdr("Цена, р."),
        _hdr("Ссылка", TA_LEFT),
    ]

    table_data = [header_row]
    row_styles = [
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_HEADER),
        ("ROWBACKGROUNDS", (0, 0), (-1, 0), [COLOR_HEADER]),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        # Center-align qty (col 2), max_price (col 3), price (col 5)
        ("ALIGN", (2, 0), (2, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("ALIGN", (5, 0), (5, -1), "CENTER"),
    ]

    r = 1  # current row index
    for item in items:
        item_id = item["id"]
        offers = results_map.get(item_id, [])

        fn_bold = body_font + "-Bold" if body_font != "Helvetica" else "Helvetica-Bold"

        # Item separator row
        item_label = f"#{item['num']}  {item['name']}"
        sep_row = [
            Paragraph(item_label, ParagraphStyle("s", fontName=fn_bold, fontSize=8, textColor=WHITE)),
            "", "", "", "", "", ""
        ]
        table_data.append(sep_row)
        row_styles.append(("BACKGROUND", (0, r), (-1, r), COLOR_SUBHDR))
        row_styles.append(("SPAN", (0, r), (-1, r)))
        r += 1

        # Description row
        description = item.get("description", "") or ""
        if description:
            desc_style = ParagraphStyle("desc", fontName=body_font, fontSize=7, leading=9,
                                        textColor=colors.HexColor("#2D5A8E"), italic=True)
            desc_row = [Paragraph(description, desc_style), "", "", "", "", "", ""]
            table_data.append(desc_row)
            row_styles.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor("#EEF4FB")))
            row_styles.append(("SPAN", (0, r), (-1, r)))
            row_styles.append(("TOPPADDING", (0, r), (-1, r), 2))
            row_styles.append(("BOTTOMPADDING", (0, r), (-1, r), 2))
            r += 1

        if not offers:
            no_row = [
                Paragraph("", small_style),
                Paragraph("— не найдено —", small_style),
                Paragraph(str(_fmt_qty_num(item["qty"])), small_style),
                Paragraph(_fmt_price(item["max_price"]) if item.get("max_price") else "—", small_style),
                "", "", ""
            ]
            table_data.append(no_row)
            row_styles.append(("BACKGROUND", (0, r), (-1, r), COLOR_NOTFOUND))
            r += 1
            continue

        for i, offer in enumerate(offers):
            price = float(offer.get("price", 0) or 0)
            is_best = (i == 0 and price > 0)
            bg = COLOR_BEST if is_best else (COLOR_FOUND if price > 0 else COLOR_NOPRICE)

            url = offer.get("url", "")
            title = offer.get("title", "") or url
            url_text = f'<link href="{url}">{title[:70]}</link>' if url else title[:70]

            offer_row = [
                Paragraph("★" if is_best else f"{i+1}.", small_style),
                Paragraph("", small_style),
                Paragraph(str(_fmt_qty_num(item["qty"])) if i == 0 else "", small_style),
                Paragraph(_fmt_price(item["max_price"]) if (i == 0 and item.get("max_price")) else "", small_style),
                Paragraph(offer.get("supplier", ""), small_style),
                Paragraph(f"<b>{price:,.0f}</b>".replace(",", " ") if price else "н/д",
                          bold_style if is_best else normal_style),
                Paragraph(url_text, link_style),
            ]
            table_data.append(offer_row)
            row_styles.append(("BACKGROUND", (0, r), (-1, r), bg))
            r += 1

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(row_styles))
    story.append(tbl)

    doc.build(story)


def _register_fonts():
    """Register DejaVu fonts for Cyrillic support if available."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
    ]
    bold_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
    ]
    for path in font_paths:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("DejaVu", path))
            except Exception:
                pass
            break
    for path in bold_paths:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("DejaVu-Bold", path))
            except Exception:
                pass
            break


def _font_available(name: str) -> bool:
    from reportlab.pdfbase import pdfmetrics
    try:
        pdfmetrics.getFont(name)
        return True
    except Exception:
        return False
